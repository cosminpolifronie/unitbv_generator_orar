# versiune: 1.4

import openpyxl
import os
import sys
import xlsxwriter


# setari
__coord_cod_orar = 'E1'
__coord_an_universitar = 'J3'
__col_an = 'A'
__col_spec = 'B'
__col_grupa = 'C'
__col_inceput_cursuri = 'E'
__font = 'Calibri'
__marime_font = 30
__culoare_border = '#9B9B9B'


# variabile
__header_time = ('8:00 – 9:50', '10:00 – 11:50', '12:00 – 13:50', '14:00 – 15:50', '16:00 – 17:50', '18:00 – 19:50', '20:00 – 21:50')
__header_day = ('Luni', 'Marți', 'Miercuri', 'Joi', 'Vineri', 'Sâmbătă')


# dictionar cu valori tuple
# {numeprescurtat: (nume intreg, culoare)}
__disciplines = {}
__ignored_disciplines = []
__professors = {}


# intoarce un array cu 4 intrari
# C/(S)/[L]
# sala
# numele materiei
# profesor
def transform_cell_value_in_formatted_array(value):
    split_value = str(value).replace(' ', '').split(',')
    split_value[0], split_value[1] = split_value[1], split_value[0]
    split_value[1], split_value[2] = split_value[2], split_value[1]
    if split_value[0].upper() == 'S':
        split_value[0] = '(S)'
    elif split_value[0].upper() == 'L':
        split_value[0] = '[L]'
    return split_value


def get_discipline_name(discipline):
    if discipline in __disciplines:
        return __disciplines[discipline][0]
    return discipline


def get_professor_name(professor):
    if professor in __professors:
        return __professors[professor]
    return professor


def get_discipline_color(discipline):
    if discipline in __disciplines:
        return __disciplines[discipline][1]
    return '#FFFFFF'


def get_col_merged_cell_value(source, coords):
    for range in source.merged_cell_ranges:
        merged_cells = list(openpyxl.utils.rows_from_range(range))
        for row in merged_cells:
            if coords in row:
                return source[merged_cells[0][0]].value
    return source[coords].value


def get_workbook_cell_format_with_color(workbook, color):
    return workbook.add_format({
            'font_name': __font,
            'font_size': __marime_font,
            'bold': False,
            'italic': False,
            'align': 'center',
            'valign': 'vcenter',
			'text_wrap': True,
            'border': 5,
            'border_color': __culoare_border,
            'bg_color': color
        })


# 1 indexed
def column_letters_to_integer(column):
    sum = 0
    for char in column:
        sum *= 26
        sum += ord(char) - ord('A') + 1
    return sum


def generate_worksheet(worksheet, source, row, version):
    # format pagina: ANSI E (44 inch x 34 inch) landscape
    worksheet.set_landscape()
    worksheet.set_paper(26)

    # inaltime header ora: 1.10 inch
    # latime header zi: 3 inch
    # inaltime camp: 5 inch (2.5 inch per rand, un camp fiind format din 2
    # randuri combinate pentru a permite afisarea materiilor din zile
    # impare/pare)
    # latime camp: 5.62 inch (2.81 inch per coloana, aceeasi poveste ca mai
    # sus, pentru a permite afisarea materiilor care se desfasoara in acelasi
    # timp)
    # inaltimea e in puncte (1 punct = 1/72 inch)
    # latimea e in numarul de caractere care incap in acel camp folosind fontul
    # standard
    worksheet.set_column(0, len(__header_time) * 2, 35)
    worksheet.set_row(0, 170)
    for i in range(1, len(__header_day) * 2 + 1):
        worksheet.set_row(i, 180)

    # generam headerele orarului
    # interval: luni-sambata, 8:00-21:50
    # populam orele si legam casutele headerului 2 cate 2
    for i in range(0, len(__header_time)):
        worksheet.merge_range(0, 2 * i + 1, 0, 2 * i + 2, __header_time[i], __bold_format)
    # populam zilele si legam casutele headerului 2 cate 2
    for i in range(0, len(__header_day)):
        worksheet.merge_range(2 * i + 1, 0, 2 * i + 2, 0, __header_day[i], __bold_format)

    # populam campul A1 cu detalii importante
    # cod orar – versiune
    # an universitar
    # an – specializare – grupa
    worksheet.write_string(0, 0, str(get_col_merged_cell_value(source, __coord_cod_orar)).replace(' ', '') + ' – ' + str(version).replace(' ', '') + '\n' + str(get_col_merged_cell_value(source, __coord_an_universitar)).replace(' ', '') + '\n' + 'A' + str(get_col_merged_cell_value(source, __col_an + str(row))) + ' – ' + str(get_col_merged_cell_value(source, __col_spec + str(row))) + '\n' + 'G – ' + str(get_col_merged_cell_value(source, __col_grupa + str(row))).replace(' ', ''), __bold_format)

    # populam campurile cu continut
    # citim cate o coloana per pas (4 casute, 1 pt.  saptamana para/impara)
    for day in range(0, len(__header_day)):
        for period in range(0, len(__header_time)):
            current_col = column_letters_to_integer(__col_inceput_cursuri) + period + day * len(__header_time)
            for col in source.iter_cols(current_col, current_col, row, row + 3):
                # generam un vector de valori
                values = []
                for item in col:
                    values.append(item.value)

                # variabila pentru a imi putea da seama daca spatiul de la
                # saptamana impara a fost umplut
                # folosita pt.  a uni toate campurile in cazul in care raman
                # goale (materia este ignorata)
                empty0 = False

                # daca ultimele 3 celule sunt goale
                if values[1] == values[2] == values[3] == None:
                    # daca si prima celula e goala
                    if values[0] == None:
                        # unim toate celulele intr-una singura
                        worksheet.merge_range(1 + day * 2, 1 + period * 2, 1 + day * 2 + 1, 1 + period * 2 + 1, '', __cell_format)
                    # avem o disciplina unica
                    else:
                        content = transform_cell_value_in_formatted_array(values[0])
                        if content[2] not in __ignored_disciplines:
                            format = get_workbook_cell_format_with_color(workbook, get_discipline_color(content[2]))
                            worksheet.merge_range(1 + day * 2, 1 + period * 2, 1 + day * 2 + 1, 1 + period * 2 + 1, '', format)
                            worksheet.write_rich_string(1 + day * 2, 1 + period * 2, __bold_text_format, content[0] + '\n', __italic_text_format, content[1] + '\n', __bold_text_format, get_discipline_name(content[2]) + '\n', __text_format, get_professor_name(content[3]), format)
                        else:
                            worksheet.merge_range(1 + day * 2, 1 + period * 2, 1 + day * 2 + 1, 1 + period * 2 + 1, '', __cell_format)
                            empty0 = True
                else:
                    # daca materiile din saptamana impara sunt identice
                    if values[0] == values[2]:
                        # daca materiile sunt nule, atunci unim campurile si nu
                        # scriem nimic
                        if values[0] == None:
                            worksheet.merge_range(1 + day * 2, 1 + period * 2, 1 + day * 2, 1 + period * 2 + 1, '', __cell_format)
                            empty0 = True
                        else:
                            content = transform_cell_value_in_formatted_array(values[0])
                            if content[2] not in __ignored_disciplines:
                                format = get_workbook_cell_format_with_color(workbook, get_discipline_color(content[2]))
                                worksheet.merge_range(1 + day * 2, 1 + period * 2, 1 + day * 2, 1 + period * 2 + 1, '', format)
                                worksheet.write_rich_string(1 + day * 2, 1 + period * 2, __bold_text_format, content[0] + '\n', __italic_text_format, content[1] + '\n', __bold_text_format, get_discipline_name(content[2]) + '\n', __text_format, get_professor_name(content[3]), format)
                            else:
                                worksheet.merge_range(1 + day * 2, 1 + period * 2, 1 + day * 2, 1 + period * 2 + 1, '', __cell_format)
                                empty0 = True
                    # daca materiile din saptamana impara difera
                    else:
                        # daca prima materie e nula, inseamna ca a doua sigur
                        # nu e
                        if values[0] == None:
                            content = transform_cell_value_in_formatted_array(values[2])
                            if content[2] not in __ignored_disciplines:
                                format = get_workbook_cell_format_with_color(workbook, get_discipline_color(content[2]))
                                worksheet.merge_range(1 + day * 2, 1 + period * 2, 1 + day * 2, 1 + period * 2 + 1, '', format)
                                worksheet.write_rich_string(1 + day * 2, 1 + period * 2, __bold_text_format, content[0] + '\n', __italic_text_format, content[1] + '\n', __bold_text_format, get_discipline_name(content[2]) + '\n', __text_format, get_professor_name(content[3]), format)
                            else:
                                worksheet.merge_range(1 + day * 2, 1 + period * 2, 1 + day * 2, 1 + period * 2 + 1, '', __cell_format)
                                empty0 = True
                        # daca a doua materie e nula, inseaman ca prima sigur
                        # nu e
                        elif values[2] == None:
                            content = transform_cell_value_in_formatted_array(values[0])
                            if content[2] not in __ignored_disciplines:
                                format = get_workbook_cell_format_with_color(workbook, get_discipline_color(content[2]))
                                worksheet.merge_range(1 + day * 2, 1 + period * 2, 1 + day * 2, 1 + period * 2 + 1, '', format)
                                worksheet.write_rich_string(1 + day * 2, 1 + period * 2, __bold_text_format, content[0] + '\n', __italic_text_format, content[1] + '\n', __bold_text_format, get_discipline_name(content[2]) + '\n', __text_format, get_professor_name(content[3]), format)
                            else:
                                worksheet.merge_range(1 + day * 2, 1 + period * 2, 1 + day * 2, 1 + period * 2 + 1, '', __cell_format)
                                empty0 = True
                        # daca ambele materii exista si sunt diferite, le
                        # scriem pe ambele
                        else:
                            content0 = transform_cell_value_in_formatted_array(values[0])
                            content2 = transform_cell_value_in_formatted_array(values[2])
                            format0 = get_workbook_cell_format_with_color(workbook, get_discipline_color(content0[2]))
                            format2 = get_workbook_cell_format_with_color(workbook, get_discipline_color(content2[2]))
                            if content0[2] not in __ignored_disciplines:
                                # daca niciuna din discipline nu e ignorata
                                if content2[2] not in __ignored_disciplines:
                                    worksheet.write_rich_string(1 + day * 2, 1 + period * 2, __bold_text_format, content0[0] + '\n', __italic_text_format, content0[1] + '\n', __bold_text_format, get_discipline_name(content0[2]) + '\n', __text_format, get_professor_name(content0[3]), format0)
                                    worksheet.write_rich_string(1 + day * 2, 1 + period * 2 + 1, __bold_text_format, content2[0] + '\n', __italic_text_format, content2[1] + '\n', __bold_text_format, get_discipline_name(content2[2]) + '\n', __text_format, get_professor_name(content2[3]), format2)
                                # daca a doua disciplina e ignorata, o scriem
                                # doar pe prima
                                else:
                                    worksheet.merge_range(1 + day * 2, 1 + period * 2, 1 + day * 2, 1 + period * 2 + 1, '', format0)
                                    worksheet.write_rich_string(1 + day * 2, 1 + period * 2, __bold_text_format, content0[0] + '\n', __italic_text_format, content0[1] + '\n', __bold_text_format, get_discipline_name(content0[2]) + '\n', __text_format, get_professor_name(content0[3]), format0)
                            # daca prima disciplina e ignorata, o scriem doar
                            # pe a doua
                            else:
                                if content2[2] not in __ignored_disciplines:
                                    worksheet.merge_range(1 + day * 2, 1 + period * 2, 1 + day * 2, 1 + period * 2 + 1, '', format2)
                                    worksheet.write_rich_string(1 + day * 2, 1 + period * 2, __bold_text_format, content2[0] + '\n', __italic_text_format, content2[1] + '\n', __bold_text_format, get_discipline_name(content2[2]) + '\n', __text_format, get_professor_name(content2[3]), format2)
                                # daca si a doua disciplina e ignorata, unim
                                # celulele si nu scriem nimic
                                else:
                                    worksheet.merge_range(1 + day * 2, 1 + period * 2, 1 + day * 2, 1 + period * 2 + 1, '', __cell_format)
                                    empty0 = True

                    # daca materiile din saptamana para sunt identice
                    if values[1] == values[3]:
                        # daca materiile sunt nule, atunci unim campurile si nu
                        # scriem nimic
                        if values[1] == None:
                            # unim toate celulele in cazul in care nu avem
                            # nimic scris in saptamana impara
                            if empty0 == True:
                                worksheet.merge_range(1 + day * 2, 1 + period * 2, 1 + day * 2 + 1, 1 + period * 2 + 1, '', __cell_format)
                            else:
                                worksheet.merge_range(1 + day * 2 + 1, 1 + period * 2, 1 + day * 2 + 1, 1 + period * 2 + 1, '', __cell_format)
                        else:
                            content = transform_cell_value_in_formatted_array(values[1])
                            if content[2] not in __ignored_disciplines:
                                format = get_workbook_cell_format_with_color(workbook, get_discipline_color(content[2]))
                                worksheet.merge_range(1 + day * 2 + 1, 1 + period * 2, 1 + day * 2 + 1, 1 + period * 2 + 1, '', format)
                                worksheet.write_rich_string(1 + day * 2 + 1, 1 + period * 2, __bold_text_format, content[0] + '\n', __italic_text_format, content[1] + '\n', __bold_text_format, get_discipline_name(content[2]) + '\n', __text_format, get_professor_name(content[3]), format)
                            else:
                                # unim toate celulele in cazul in care nu avem
                                # nimic scris in saptamana impara
                                if empty0 == True:
                                    worksheet.merge_range(1 + day * 2, 1 + period * 2, 1 + day * 2 + 1, 1 + period * 2 + 1, '', __cell_format)
                                else:
                                    worksheet.merge_range(1 + day * 2 + 1, 1 + period * 2, 1 + day * 2 + 1, 1 + period * 2 + 1, '', __cell_format)
                    # daca materiile din saptamana para difera
                    else:
                        # daca prima materie e nula, inseamna ca a doua sigur
                        # nu e
                        if values[1] == None:
                            content = transform_cell_value_in_formatted_array(values[3])
                            if content[2] not in __ignored_disciplines:
                                format = get_workbook_cell_format_with_color(workbook, get_discipline_color(content[2]))
                                worksheet.merge_range(1 + day * 2 + 1, 1 + period * 2, 1 + day * 2 + 1, 1 + period * 2 + 1, '', format)
                                worksheet.write_rich_string(1 + day * 2 + 1, 1 + period * 2, __bold_text_format, content[0] + '\n', __italic_text_format, content[1] + '\n', __bold_text_format, get_discipline_name(content[2]) + '\n', __text_format, get_professor_name(content[3]), format)
                            else:
                                # unim toate celulele in cazul in care nu avem
                                # nimic scris in saptamana impara
                                if empty0 == True:
                                    worksheet.merge_range(1 + day * 2, 1 + period * 2, 1 + day * 2 + 1, 1 + period * 2 + 1, '', __cell_format)
                                else:
                                    worksheet.merge_range(1 + day * 2 + 1, 1 + period * 2, 1 + day * 2 + 1, 1 + period * 2 + 1, '', __cell_format)
                        # daca a doua materie e nula, inseaman ca prima sigur
                        # nu e
                        elif values[3] == None:
                            content = transform_cell_value_in_formatted_array(values[1])
                            if content[2] not in __ignored_disciplines:
                                format = get_workbook_cell_format_with_color(workbook, get_discipline_color(content[2]))
                                worksheet.merge_range(1 + day * 2 + 1, 1 + period * 2, 1 + day * 2 + 1, 1 + period * 2 + 1, '', format)
                                worksheet.write_rich_string(1 + day * 2 + 1, 1 + period * 2, __bold_text_format, content[0] + '\n', __italic_text_format, content[1] + '\n', __bold_text_format, get_discipline_name(content[2]) + '\n', __text_format, get_professor_name(content[3]), format)
                            else:
                                # unim toate celulele in cazul in care nu avem
                                # nimic scris in saptamana impara
                                if empty0 == True:
                                    worksheet.merge_range(1 + day * 2, 1 + period * 2, 1 + day * 2 + 1, 1 + period * 2 + 1, '', __cell_format)
                                else:
                                    worksheet.merge_range(1 + day * 2 + 1, 1 + period * 2, 1 + day * 2 + 1, 1 + period * 2 + 1, '', __cell_format)
                        # daca ambele materii exista si sunt diferite, le
                        # scriem pe ambele
                        else:
                            content1 = transform_cell_value_in_formatted_array(values[1])
                            content3 = transform_cell_value_in_formatted_array(values[3])
                            format1 = get_workbook_cell_format_with_color(workbook, get_discipline_color(content1[2]))
                            format3 = get_workbook_cell_format_with_color(workbook, get_discipline_color(content3[2]))
                            if content1[2] not in __ignored_disciplines:
                                # daca niciuna din discipline nu e ignorata
                                if content3[2] not in __ignored_disciplines:
                                    worksheet.write_rich_string(1 + day * 2 + 1, 1 + period * 2, __bold_text_format, content1[0] + '\n', __italic_text_format, content1[1] + '\n', __bold_text_format, get_discipline_name(content1[2]) + '\n', __text_format, get_professor_name(content1[3]), format1)
                                    worksheet.write_rich_string(1 + day * 2 + 1, 1 + period * 2 + 1, __bold_text_format, content3[0] + '\n', __italic_text_format, content3[1] + '\n', __bold_text_format, get_discipline_name(content3[2]) + '\n', __text_format, get_professor_name(content3[3]), format3)
                                # daca a doua disciplina e ignorata, o scriem
                                # doar pe prima
                                else:
                                    worksheet.merge_range(1 + day * 2 + 1, 1 + period * 2, 1 + day * 2 + 1, 1 + period * 2 + 1, '', format1)
                                    worksheet.write_rich_string(1 + day * 2 + 1, 1 + period * 2, __bold_text_format, content1[0] + '\n', __italic_text_format, content1[1] + '\n', __bold_text_format, get_discipline_name(content1[2]) + '\n', __text_format, get_professor_name(content1[3]), format1)
                            # daca prima disciplina e ignorata, o scriem doar
                            # pe a doua
                            else:
                                if content3[2] not in __ignored_disciplines:
                                    worksheet.merge_range(1 + day * 2 + 1, 1 + period * 2, 1 + day * 2 + 1, 1 + period * 2 + 1, '', format3)
                                    worksheet.write_rich_string(1 + day * 2 + 1, 1 + period * 2, __bold_text_format, content3[0] + '\n', __italic_text_format, content3[1] + '\n', __bold_text_format, get_discipline_name(content3[2]) + '\n', __text_format, get_professor_name(content3[3]), format3)
                                # daca si a doua disciplina e ignorata, unim
                                # celulele si nu scriem nimic
                                else:
                                    # unim toate celulele in cazul in care nu
                                    # avem nimic scris in saptamana impara
                                    if empty0 == True:
                                        worksheet.merge_range(1 + day * 2, 1 + period * 2, 1 + day * 2 + 1, 1 + period * 2 + 1, '', __cell_format)
                                    else:
                                        worksheet.merge_range(1 + day * 2 + 1, 1 + period * 2, 1 + day * 2 + 1, 1 + period * 2 + 1, '', __cell_format)


if __name__ == "__main__":
    if len(sys.argv) < 3:
        sys.exit('Mai putin de 3 argumente!\nFolosire: unitbv_generator_orar.py\n\tcale_orar\n\tfolder_output\n\tprimul_rand_al_grupei_de_generat\n\tprimul_rand_al_grupei_de_generat\n\tetc.')
    
    current_dir = os.path.dirname(os.path.realpath(__file__))
	
	# incarcam fisierul cu materii
    temp_file = open(current_dir + '\materii.txt', 'r', encoding='utf-8')
    lines = temp_file.readlines()
    for line in lines:
        data = line.split('=')
        if len(data) == 3:
            __disciplines[data[0]] = (data[1], data[2].replace('\n', ''))
    temp_file.close()
            
    # incarcam fisierul cu profesori
    temp_file = open(current_dir + '\profesori.txt', 'r', encoding='utf-8')
    lines = temp_file.readlines()
    for line in lines:
        data = line.split('=')
        if len(data) == 2:
            __professors[data[0]] = data[1].replace('\n', '')
    temp_file.close()

    # incarcam fisierul cu materii ignorate
    temp_file = open(current_dir + '\materii_ignorate.txt', 'r', encoding='utf-8')
    lines = temp_file.readlines()
    for line in lines:
         __ignored_disciplines.append(line.replace('\n', ''))
    temp_file.close()
	
	# extragem versiunea orarului
    version = sys.argv[1].split('-')
    version = version[len(version) - 1].split('.')[0].replace(' ', '')

    # setam sursa orarului
    source = openpyxl.load_workbook(sys.argv[1]).active
        
    # generam workbook-ul curent
    workbook = xlsxwriter.Workbook(sys.argv[2] + '\\orar-' + str(get_col_merged_cell_value(source, __coord_cod_orar)).replace(' ', '') + '-' + version + '.xlsx')

    __text_format = workbook.add_format({
            'font_name': __font,
            'font_size': __marime_font,
            'bold': False,
            'italic': False
        })

    __bold_text_format = workbook.add_format({
            'font_name': __font,
            'font_size': __marime_font,
            'bold': True,
            'italic': False
        })

    __italic_text_format = workbook.add_format({
            'font_name': __font,
            'font_size': __marime_font,
            'bold': False,
            'italic': True
        })

    __bold_format = workbook.add_format({
            'font_name': __font,
            'font_size': __marime_font,
            'bold': True,
            'italic': False,
            'align': 'center',
            'valign': 'vcenter',
			'text_wrap': True,
            'border': 5,
            'border_color': __culoare_border
        })

    __cell_format = workbook.add_format({
            'font_name': __font,
            'font_size': __marime_font,
            'bold': False,
            'italic': False,
            'align': 'center',
            'valign': 'vcenter',
			'text_wrap': True,
            'border': 5,
            'border_color': __culoare_border
        })

    # pentru fiecare rand oferit ca parametru generam un nou worksheet in
    # fisier
    for row in sys.argv[3:]:
        worksheet = workbook.add_worksheet(str(str(get_col_merged_cell_value(source, __col_an + str(row))) + '-' + str(get_col_merged_cell_value(source, __col_spec + str(row))) + '-' + str(get_col_merged_cell_value(source, __col_grupa + str(row)))).replace(' ', ''))
        generate_worksheet(worksheet, source, int(row), version)
    
    # salvam fisierul
    workbook.close()
